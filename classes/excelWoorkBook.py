from openpyxl import load_workbook
from tkinter import messagebox

class excelWorkBook():
    def ___init___(self):
        self.wsListaTurno = None
    
    def initWorkBook(self, path):
        try:
            self.wbListaTurno = load_workbook(path, data_only = True)
        except:
            return False
        return True
    
    def getSheets(self):
        return self.wbListaTurno.sheetnames

    def initSheet(self, sheet):
        self.wsListaTurno = self.wbListaTurno[sheet]

    def process(self, pathFileSave, dayFrom, dayTo):

        dayMonth = self.daysMax()

        dayFrom = self.dayFrom(dayMonth, dayFrom)
        if not dayFrom:
            messagebox.showerror('Error de Periodo', 'No ingreso correctamente el periodo ("dayFrom" fuera de rango)')
            return
        
        dayTo = self.dayTo(dayMonth, dayFrom, dayTo)
        if not dayTo:
            messagebox.showerror('Error de Periodo', 'No ingreso correctamente el periodo ("dayTo" fuera de rango)')
            return

        self.initRangeList()

        #self.tecSelfRange.sorted()
        self.getValue()
        self.tecSelfList = sorted(self.tecSelfList , key=lambda x: x[0])

        for column in self.tecSelfList:
            print(' ', end= '\n')
            for value in column:
                print(str(value) + ' ', end= '')

        self.initTemplate()

        #print(self.tecSelfRange[0][0].value)
        
        self.saveInTemplate(dayTo)
        self.saveWorkBook(pathFileSave, dayFrom, dayTo)
        print("cierre for")
        messagebox.showinfo("Proceso completado", "Se guardaron los partes con exito")
        
    
    def saveWorkBook(self, pathFileSave, dayFrom, dayTo):
        meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]  

        i=0
        while self.wsListaTurno["B7"].value.upper() != meses[i] or i > 12:
            i += 1

        if i < 10:
            mes = "0" + str(i + 1)
        else:
            mes = str(i + 1)

        self.wsRacionamiento["C" + str(10)].value = meses[i] + " de " + str(self.wsListaTurno["H7"].value)
        #print(pathFileSave + "/Parte día " + dia + '-' + mes + '-' + str(self.wsListaTurno["H7"].value))
        self.wbRacionamiento.save(pathFileSave + "/Parte del mes de " + meses[i] + ' del ' + str(self.wsListaTurno["H7"].value) + ".xlsx")
    
    def initRangeList(self):
        listTitle = 3 
        cellToDelete = 0
        listStart = 13
        listEnd = listStart
        
        while (self.wsListaTurno['A' + str(listEnd)].value):
            listEnd += 1
        

        while(self.wsListaTurno['A' + str(listEnd + cellToDelete)].value != self.wsListaTurno["H7"].value):
            cellToDelete += 1
        
        print(str(cellToDelete))

        self.wsListaTurno.delete_rows(listEnd, cellToDelete + 1)
        cellToDelete = 0

        print('primer delete' + str(cellToDelete + 1))
        print( self.wsListaTurno['A' + str(listEnd)].value)

        while (self.wsListaTurno['A' + str(listEnd)].value):
            listEnd += 1
        
        while(self.wsListaTurno['A' + str(listEnd + cellToDelete)].value != None):
            cellToDelete += 1

        print(str(cellToDelete))


        self.wsListaTurno.delete_rows(listEnd, cellToDelete + 1)

        print('segundo delete' + str(cellToDelete))
        print( self.wsListaTurno['A' + str(listEnd)].value)

        while (self.wsListaTurno['A' + str(listEnd)].value):
            listEnd += 1

        self.tecSelfRange = self.wsListaTurno['A' + str(listStart) : 'AF' + str(listEnd - 1)]
        
        print( self.wsListaTurno['A' + str(listEnd)].value)

        for i in range(13, listEnd):
            print(self.wsListaTurno['A' + str(i)].value)


        print(" ")
        print(" ")

        print(self.tecSelfRange)

        print(" ")
        print(" ")


        #for row in self.tecSelfRange:
         #   print(' ', end= '\n')
          #  for cell in row:
           #     print(str(cell.value) + ' ', end= '')
        print(" ")
        print(" ")

        self.tecNumber = listEnd - listStart + 1

        #self.tecSelfList = [[] for i in range(listEnd - listStart + 1)]

        #for row in self.tecSelfList:
        #   for i in range(32):
         #       row.append(None)

        #self.tecSelfList = [listEnd - listStart + 1][32]

    def dayFrom(self, daysMonth, dayFrom):
        if not dayFrom:
            dayFrom = 1
            print("dayFrom = 1")
        else:
            dayFrom = int(dayFrom)
            print(dayFrom, daysMonth)
            if (dayFrom < 1 or dayFrom > daysMonth):
                return False
        
        return dayFrom
    
    def dayTo(self, daysMonth, dayFrom, dayTo):
        if not dayTo:
            dayTo = daysMonth
            print("dayTo = " + str(daysMonth))
        else:
            dayTo = int(dayTo)
            print("dayTo = ", dayTo)
            print("daysMonth = ", daysMonth)
            if(dayTo < 1 or dayTo > daysMonth or dayTo < dayFrom):
                return False
        
        return dayTo

    def daysMax(self):
        daysMonth = 1
        cell_RangeListaTurno = self.wsListaTurno['B12':'AF12'][0]
        for cell in cell_RangeListaTurno:
            if type(cell.value) is int:
                daysMonth += 1
        daysMonth -= 1
        print("daysMont = " + str(daysMonth))

        return daysMonth

    def initTemplate(self,):
        self.wbRacionamiento = load_workbook('resourse/excel/plantilla.xlsx')
        self.wsRacionamiento = self.wbRacionamiento["Racionamiento"]

        iniListRa = 12
        finListRa = self.tecNumber + 12

        print(finListRa - iniListRa)

        self.cellRangeRaTec = self.wsRacionamiento['C' + str(iniListRa) : 'AH' + str(finListRa)]

    def searchTech(self, technical, cantTec):
        pos = 0
        if technical == "TORRES Jose":
            technical = "TORRES FERN"
        print("buscarTec")
        
        while pos < cantTec and self.cellRangeRaTec[pos][0] and self.cellRangeRaTec[pos][0].value[0:8] != technical[0:8]:
            print( self.cellRangeRaTec[pos][0].value[0:8] + "==" + technical[0:8])
            pos += 1
        
        if pos == cantTec:
            return -1
        return pos
        
    def celdProcess(self, cellRange, technical, row, cellColum, pos, type):
        print("procesarCelda")
        print(str(technical) + str(cellRange[row][cellColum].value) + "------" + str(row) + str(cellColum))

        if cellColum > 0 and cellRange[row][cellColum - 1].value == 'TM' or cellRange[row][cellColum - 1].value == 'TT':
            self.cellRangeRa[pos][0].value = "Descanso de turno"

        if cellRange[row][cellColum].value == 'TM':
            self.cellRangeRa[pos][0].value = type
            self.cellRangeRa[pos][1].value = "12 Hs Mañana"
                        
        elif cellRange[row][cellColum].value == 'TT':
            self.cellRangeRa[pos][0].value = type
            self.cellRangeRa[pos][1].value = "12 Hs Tarde"
                    
        elif cellRange[row][cellColum].value == 'L':
            self.cellRangeRa[pos][0].value = "Licencia"
            self.cellRangeRa[pos][1].value = "Fecha de"
            i=0
            Ldesde = None
            Lhasta = None

            while self.cellRangeLicen[i][0].value and self.cellRangeLicen[i][0].value != technical and i<10:
                i += 1

            if self.cellRangeLicen[i][0].value == technical:
                Ldesde = self.cellRangeLicen[i][2].value
                Lhasta = self.cellRangeLicen[i][5].value

            if Ldesde and Lhasta:
                self.cellRangeRa[pos][2].value = str(Ldesde.strftime('%d/%m/%Y')) + " al " + str(Lhasta.strftime('%d/%m/%Y'))
            else:
                return False            

        elif cellRange[row][cellColum].value == 'P':
            self.cellRangeRa[pos][0].value = "Presente"
            self.cellRangeRa[pos][1].value = "08 Hs"
                    
        elif cellRange[row][cellColum].value == 'C':
            self.cellRangeRa[pos][0].value = "Comisión"
            self.cellRangeRa[pos][1].value = "Fecha de"
            i=0
            Cdesde = None
            Chasta = None
            print("entro com")
            while self.cellRangeCom[i][8].value and self.cellRangeCom[i][8].value!= technical and i<10:
                i += 1
                
            if self.cellRangeCom[i][8].value == technical:
                Cdesde = self.cellRangeCom[i][2].value
                Chasta = self.cellRangeCom[i][5].value

            if Cdesde and Chasta:
                self.cellRangeRa[pos][2].value = str(Cdesde.strftime('%d/%m/%Y')) + " al " + str(Chasta.strftime('%d/%m/%Y'))
            else:
                return False

        elif cellRange[row][cellColum].value == 'PE':
                self.cellRangeRa[pos][1].value = "Parte de enfermo"
                    
        elif cellRange[row][cellColum].value == 'CU':
                self.cellRangeRa[pos][0].value = "Curso"

        elif not self.cellRangeRa[pos][0].value:
                self.cellRangeRa[pos][0].value = "Franco" 

        return True     

    def getValue(self):
        i = 0
        rowList = []
        self.tecSelfList = []
        for row in self.tecSelfRange:
            for cell in row:
                rowList.append(cell.value)
                #print(rowList)
            self.tecSelfList.append(rowList)
            rowList = []
    #    for column in self.tecSelfList:
     #       print("")
      #      for value in column:
       #         print(value)

   # def sortRange():
    #    minName = self.tecSelfList[0][1]
     #   i = 0
      #  for name in self.tecSelfList[][0]
       #     if name < minName
        #        for row in self.tecSelfList[]:
          #          for value in row:
         #               aux = value

    def saveInTemplate(self, dayTo):
        i=0
        j=0 
        lastThec = ''
        for row in self.tecSelfList:
            if lastThec == self.tecSelfList[j][i]:
                j -= 1
            for cell in row:
                #print(cell, end='')
                #print(j, end='')
                #print()
                if cell == None:
                    self.cellRangeRaTec[j][i].value = 'F'
                else:
                    self.cellRangeRaTec[j][i].value = cell
                i += 1
                if(i == dayTo+1):
                    break

            lastThec = self.cellRangeRaTec[j][0].value    
            i = 0
            j += 1
            print(str(lastThec) + '==' + str(self.tecSelfList[j][i]))

